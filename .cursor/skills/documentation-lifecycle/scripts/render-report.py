#!/usr/bin/env -S uv run --python 3.12
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///

"""Render structured report JSON into HTML or XLSX artifacts."""

from __future__ import annotations

import argparse
from datetime import datetime
import html
import json
from pathlib import Path
import re
import sys
from typing import Any

EXIT_SUCCESS = 0
EXIT_ERROR = 1

FORMAT_HTML = "html"
FORMAT_XLSX = "xlsx"
SUPPORTED_FORMATS = (FORMAT_HTML, FORMAT_XLSX)

SECTION_TYPES = {"table", "text", "keyvalue", "list"}


def fail(message: str) -> int:
    sys.stderr.write(f"{message}\n")
    return EXIT_ERROR


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="render-report.py",
        description="Render structured report JSON to HTML or XLSX.",
    )
    parser.add_argument(
        "input",
        nargs="?",
        help="Input JSON path. If omitted, read JSON from stdin.",
    )
    parser.add_argument(
        "--format",
        choices=list(SUPPORTED_FORMATS),
        default=FORMAT_HTML,
        help="Output format.",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Output path. Default uses title and timestamp.",
    )
    parser.add_argument(
        "--title",
        help="Override report title from input JSON.",
    )
    return parser.parse_args(argv)


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower()).strip("-")
    return cleaned or "report"


def make_default_output_path(title: str, fmt: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = ".html" if fmt == FORMAT_HTML else ".xlsx"
    return Path(f"{slugify(title)}-{stamp}{suffix}")


def read_input_text(input_path: str | None) -> str:
    if input_path is not None:
        path = Path(input_path)
        if not path.exists():
            raise ValueError(f"Input file not found: {path}")
        return path.read_text(encoding="utf-8")
    if sys.stdin.isatty():
        raise ValueError("No input file provided and stdin is empty.")
    return sys.stdin.read()


def parse_report_json(raw_text: str) -> dict[str, Any]:
    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Invalid JSON at line {exc.lineno}, column {exc.colno}: {exc.msg}"
        ) from exc
    if not isinstance(data, dict):
        raise ValueError("Top-level JSON value must be an object.")
    return data


def validate_sections(sections: Any) -> list[dict[str, Any]]:
    if sections is None:
        return []
    if not isinstance(sections, list):
        raise ValueError("'sections' must be an array.")
    validated: list[dict[str, Any]] = []
    for index, section in enumerate(sections):
        if not isinstance(section, dict):
            raise ValueError(f"Section {index} must be an object.")
        section_type = section.get("type")
        if section_type not in SECTION_TYPES:
            raise ValueError(
                f"Section {index} has unsupported type '{section_type}'. "
                f"Expected one of: {', '.join(sorted(SECTION_TYPES))}."
            )
        validated.append(section)
    return validated


def extract_report(data: dict[str, Any], override_title: str | None) -> dict[str, Any]:
    title = override_title or data.get("title") or "Report"
    if not isinstance(title, str) or title.strip() == "":
        raise ValueError("Report title must be a non-empty string.")
    subtitle = data.get("subtitle")
    generated_at = data.get("generated_at")
    metadata = data.get("metadata") or {}
    if subtitle is not None and not isinstance(subtitle, str):
        raise ValueError("'subtitle' must be a string when provided.")
    if generated_at is not None and not isinstance(generated_at, str):
        raise ValueError("'generated_at' must be a string when provided.")
    if not isinstance(metadata, dict):
        raise ValueError("'metadata' must be an object.")
    return {
        "title": title.strip(),
        "subtitle": subtitle,
        "generated_at": generated_at,
        "metadata": metadata,
        "sections": validate_sections(data.get("sections")),
    }


def html_table(section: dict[str, Any]) -> str:
    columns = section.get("columns")
    rows = section.get("rows")
    if not isinstance(columns, list):
        raise ValueError("Table section requires 'columns' array.")
    if not isinstance(rows, list):
        raise ValueError("Table section requires 'rows' array.")
    head_cells = "".join(f"<th>{html.escape(str(col))}</th>" for col in columns)
    body_rows: list[str] = []
    for row in rows:
        if not isinstance(row, list):
            raise ValueError("Each table row must be an array.")
        cells = "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row)
        body_rows.append(f"<tr>{cells}</tr>")
    return (
        '<div class="card"><table><thead><tr>'
        f"{head_cells}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>"
    )


def html_keyvalue(section: dict[str, Any]) -> str:
    data = section.get("data")
    if not isinstance(data, dict):
        raise ValueError("Keyvalue section requires 'data' object.")
    rows = "".join(
        "<tr>"
        f"<th>{html.escape(str(key))}</th>"
        f"<td>{html.escape(str(value))}</td>"
        "</tr>"
        for key, value in data.items()
    )
    return f'<div class="card"><table class="kv"><tbody>{rows}</tbody></table></div>'


def html_list(section: dict[str, Any]) -> str:
    items = section.get("items")
    if not isinstance(items, list):
        raise ValueError("List section requires 'items' array.")
    rendered = "".join(f"<li>{html.escape(str(item))}</li>" for item in items)
    return f'<div class="card"><ul>{rendered}</ul></div>'


def html_text(section: dict[str, Any]) -> str:
    content = section.get("content")
    if not isinstance(content, str):
        raise ValueError("Text section requires 'content' string.")
    paragraphs = "".join(
        f"<p>{html.escape(part.strip())}</p>" for part in content.split("\n") if part.strip()
    )
    return f'<div class="card">{paragraphs or "<p></p>"}</div>'


def render_html(report: dict[str, Any]) -> str:
    style = """
<style>
:root { color-scheme: light dark; --bg:#f6f8fb; --panel:#ffffff; --text:#1a1f2b; --muted:#5c6475; --border:#d8deea; --accent:#2f6feb; --row:#f1f5fb; }
@media (prefers-color-scheme: dark) { :root { --bg:#0f1320; --panel:#161c2c; --text:#e8edf6; --muted:#9eabc2; --border:#2a3550; --accent:#82a8ff; --row:#1b2337; } }
* { box-sizing: border-box; }
body { margin:0; padding:2rem; background:var(--bg); color:var(--text); font: 15px/1.5 "Segoe UI", Inter, Helvetica, Arial, sans-serif; }
.page { max-width: 1100px; margin: 0 auto; }
h1,h2,h3 { margin:0; line-height:1.2; }
h1 { font-size: 1.9rem; margin-bottom: .35rem; }
h2 { font-size: 1.2rem; margin: 1.5rem 0 .6rem; color:var(--text); }
.subtitle { color:var(--muted); margin-bottom: .8rem; }
.meta { margin: 1rem 0 1.25rem; padding: .9rem 1rem; border:1px solid var(--border); border-radius:10px; background:var(--panel); }
.meta-grid { display:grid; grid-template-columns: 220px 1fr; gap:.4rem 1rem; }
.meta-key { color:var(--muted); font-weight:600; }
.card { background:var(--panel); border:1px solid var(--border); border-radius:10px; padding:.6rem; overflow:auto; }
table { width:100%; border-collapse: collapse; font-size:.94rem; }
th, td { border:1px solid var(--border); text-align:left; padding:.55rem .65rem; vertical-align: top; }
thead th { background:color-mix(in oklab, var(--accent) 14%, var(--panel)); font-weight:700; letter-spacing:.01em; }
tbody tr:nth-child(even) td { background:var(--row); }
table.kv th { width: 35%; background: transparent; }
ul { margin:.25rem 0 .25rem 1.15rem; padding:0; }
p { margin:.35rem 0; white-space: pre-wrap; }
@media print { body { background:#fff; padding:0.5in; } .card, .meta { break-inside: avoid; } }
</style>
"""
    subtitle = report["subtitle"] or ""
    generated_label = report["generated_at"] or datetime.now().isoformat(timespec="seconds")
    metadata_rows = "".join(
        f'<div class="meta-key">{html.escape(str(key))}</div><div>{html.escape(str(value))}</div>'
        for key, value in report["metadata"].items()
    )
    sections_html: list[str] = []
    section_renderers = {
        "table": html_table,
        "text": html_text,
        "keyvalue": html_keyvalue,
        "list": html_list,
    }
    for section in report["sections"]:
        heading = html.escape(str(section.get("heading") or "Section"))
        section_type = section["type"]
        body = section_renderers.get(section_type, html_list)(section)
        sections_html.append(f"<section><h2>{heading}</h2>{body}</section>")
    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        f"<meta name='viewport' content='width=device-width, initial-scale=1'>{style}"
        f"</head><body><main class='page'><h1>{html.escape(report['title'])}</h1>"
        f"<p class='subtitle'>{html.escape(subtitle)}</p>"
        "<section class='meta'><div class='meta-grid'>"
        "<div class='meta-key'>Generated</div>"
        f"<div>{html.escape(generated_label)}</div>{metadata_rows}</div></section>"
        f"{''.join(sections_html)}</main></body></html>"
    )


def render_xlsx(report: dict[str, Any], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "XLSX rendering requires openpyxl. Run via 'uv run --script render-report.py' "
            "to auto-install script dependencies."
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise RuntimeError("Failed to initialize workbook worksheet.")
    worksheet.title = "Report"
    row = 1

    title_font = Font(size=16, bold=True)
    section_font = Font(size=12, bold=True)
    header_fill = PatternFill(fill_type="solid", start_color="D9E2F3", end_color="D9E2F3")

    worksheet.cell(row=row, column=1, value=report["title"]).font = title_font
    row += 1
    if report["subtitle"]:
        worksheet.cell(row=row, column=1, value=report["subtitle"])
        row += 1
    worksheet.cell(row=row, column=1, value="Generated")
    worksheet.cell(row=row, column=2, value=report["generated_at"] or datetime.now().isoformat())
    row += 1

    for key, value in report["metadata"].items():
        worksheet.cell(row=row, column=1, value=str(key))
        worksheet.cell(row=row, column=2, value=str(value))
        row += 1
    row += 1

    table_counter = 1
    for section in report["sections"]:
        heading = str(section.get("heading") or "Section")
        worksheet.cell(row=row, column=1, value=heading).font = section_font
        row += 1

        section_type = section["type"]
        if section_type == "table":
            columns = section.get("columns")
            rows = section.get("rows")
            if not isinstance(columns, list) or not isinstance(rows, list):
                raise ValueError("Table section requires 'columns' and 'rows' arrays.")
            start = row
            for col_idx, name in enumerate(columns, start=1):
                cell = worksheet.cell(row=row, column=col_idx, value=str(name))
                cell.font = Font(bold=True)
                cell.fill = header_fill
            row += 1
            for data_row in rows:
                if not isinstance(data_row, list):
                    raise ValueError("Each table row must be an array.")
                for col_idx, value in enumerate(data_row, start=1):
                    worksheet.cell(row=row, column=col_idx, value=value)
                row += 1
            end = row - 1 if rows else start
            if columns:
                start_ref = f"A{start}"
                end_ref = f"{get_column_letter(len(columns))}{end}"
                if rows:
                    table = Table(displayName=f"SectionTable{table_counter}", ref=f"{start_ref}:{end_ref}")
                    table.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    worksheet.add_table(table)
                    table_counter += 1
                else:
                    worksheet.auto_filter.ref = f"{start_ref}:{end_ref}"
            row += 1
            continue

        if section_type == "keyvalue":
            data = section.get("data")
            if not isinstance(data, dict):
                raise ValueError("Keyvalue section requires 'data' object.")
            for key, value in data.items():
                worksheet.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
                worksheet.cell(row=row, column=2, value=str(value))
                row += 1
            row += 1
            continue

        if section_type == "list":
            items = section.get("items")
            if not isinstance(items, list):
                raise ValueError("List section requires 'items' array.")
            for item in items:
                worksheet.cell(row=row, column=1, value=f"- {item}")
                row += 1
            row += 1
            continue

        content = section.get("content")
        if not isinstance(content, str):
            raise ValueError("Text section requires 'content' string.")
        worksheet.cell(row=row, column=1, value=content).alignment = Alignment(wrap_text=True)
        row += 2

    for col in range(1, 7):
        worksheet.column_dimensions[get_column_letter(col)].width = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_html(report: dict[str, Any], output_path: Path) -> None:
    html_output = render_html(report)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html_output, encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    try:
        raw_input = read_input_text(args.input)
        report_data = parse_report_json(raw_input)
        report = extract_report(report_data, args.title)
    except ValueError as exc:
        return fail(str(exc))

    output_path = Path(args.output) if args.output else make_default_output_path(report["title"], args.format)
    try:
        if args.format == FORMAT_HTML:
            write_html(report, output_path)
        else:
            render_xlsx(report, output_path)
    except (ValueError, RuntimeError) as exc:
        return fail(str(exc))
    except OSError as exc:
        return fail(f"Failed to write output: {exc}")

    print(str(output_path))
    return EXIT_SUCCESS


if __name__ == "__main__":
    raise SystemExit(main())
